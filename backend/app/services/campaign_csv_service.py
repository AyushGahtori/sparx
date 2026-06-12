import csv
import io
import re
import zipfile
from functools import lru_cache
from xml.etree import ElementTree

from fastapi import UploadFile
from openpyxl import load_workbook
from pypdf import PdfReader

from app.config.settings import Settings, get_settings
from app.core.errors import AppError
from app.schemas.campaign import (
    CampaignContactInput,
    CampaignCsvPreviewResponse,
    CampaignCsvPreviewRow,
)
from app.utils.phone import normalize_phone_number


class CampaignCsvService:
    required_columns = {"name", "phone"}
    optional_columns = {
        "company",
        "city",
        "state",
        "country",
        "role",
        "email",
        "website",
        "interest",
        "notes",
    }
    supported_columns = required_columns | optional_columns
    supported_file_types = {
        ".csv": "csv",
        ".xlsx": "xlsx",
        ".xls": "xls",
    }
    header_aliases = {
        "name": {
            "name",
            "full name",
            "lead name",
            "contact name",
            "person name",
            "customer name",
        },
        "phone": {
            "phone",
            "phone number",
            "number",
            "mobile",
            "mobile number",
            "contact number",
            "telephone",
            "telephone number",
            "tel",
        },
        "company": {
            "company",
            "company name",
            "organization",
            "organisation",
            "business",
            "account",
        },
        "city": {"city", "town"},
        "state": {"state", "province", "region"},
        "country": {"country", "nation"},
        "role": {"role", "designation", "title", "job title", "position"},
        "email": {"email", "email address", "mail"},
        "website": {"website", "company website", "url", "site"},
        "interest": {
            "interest",
            "product interest",
            "service interest",
            "requirement",
            "requirements",
            "need",
            "needs",
            "type",
            "company type",
            "business type",
            "industry",
            "sector",
            "category",
        },
        "notes": {"notes", "remarks", "comment", "comments", "description", "summary"},
    }
    text_field_patterns = {
        "name": re.compile(r"^(?:name|full name|lead name|contact name|customer name)\s*[:\-]\s*(.+)$", re.IGNORECASE),
        "phone": re.compile(
            r"^(?:phone|phone number|mobile|mobile number|contact number|telephone|tel)\s*[:\-]\s*(.+)$",
            re.IGNORECASE,
        ),
        "company": re.compile(r"^(?:company|company name|organization|organisation|business)\s*[:\-]\s*(.+)$", re.IGNORECASE),
        "city": re.compile(r"^(?:city|town)\s*[:\-]\s*(.+)$", re.IGNORECASE),
        "state": re.compile(r"^(?:state|province|region)\s*[:\-]\s*(.+)$", re.IGNORECASE),
        "country": re.compile(r"^(?:country|nation)\s*[:\-]\s*(.+)$", re.IGNORECASE),
        "role": re.compile(r"^(?:role|designation|title|job title|position)\s*[:\-]\s*(.+)$", re.IGNORECASE),
        "email": re.compile(r"^(?:email|email address|mail)\s*[:\-]\s*(.+)$", re.IGNORECASE),
        "website": re.compile(r"^(?:website|company website|url|site)\s*[:\-]\s*(.+)$", re.IGNORECASE),
        "interest": re.compile(
            r"^(?:interest|product interest|service interest|requirement|requirements|need|needs)\s*[:\-]\s*(.+)$",
            re.IGNORECASE,
        ),
        "notes": re.compile(r"^(?:notes|remarks|comment|comments|description|summary)\s*[:\-]\s*(.+)$", re.IGNORECASE),
    }

    def __init__(self, settings: Settings) -> None:
        self.settings = settings

    async def preview_upload(self, upload_file: UploadFile) -> CampaignCsvPreviewResponse:
        filename = upload_file.filename or "campaign-leads.csv"
        suffix = self._resolve_suffix(filename)
        file_type = self.supported_file_types[suffix]

        raw_content = await upload_file.read()
        if not raw_content:
            raise AppError(
                status_code=400,
                code="empty_lead_file",
                message="The uploaded lead file is empty.",
            )
        if len(raw_content) > self.settings.campaign_csv_max_file_size_bytes:
            raise AppError(
                status_code=413,
                code="lead_file_too_large",
                message=(
                    f"Lead uploads must be smaller than {self.settings.campaign_csv_max_file_size_bytes} bytes."
                ),
            )

        extracted_rows, source_columns, unmapped_columns = self._extract_rows(
            filename=filename,
            suffix=suffix,
            raw_content=raw_content,
        )
        if not extracted_rows:
            raise AppError(
                status_code=400,
                code="no_contacts_detected",
                message=(
                    "We could not detect any lead contacts in this file. "
                    "Use a table with contact columns or a document with labeled fields such as Name and Phone."
                ),
            )

        preview_rows: list[CampaignCsvPreviewRow] = []
        valid_contacts: list[CampaignContactInput] = []
        seen_phones: set[str] = set()

        for index, extracted_row in enumerate(extracted_rows, start=1):
            extracted_row = self._normalize_contact_identity(extracted_row)
            if index > self.settings.campaign_csv_max_rows:
                raise AppError(
                    status_code=400,
                    code="lead_row_limit_exceeded",
                    message=(
                        f"Lead uploads cannot contain more than {self.settings.campaign_csv_max_rows} contacts."
                    ),
                )

            preview_row = CampaignCsvPreviewRow(
                row_number=extracted_row.get("row_number") or index,
                name=extracted_row.get("name"),
                phone=extracted_row.get("phone"),
                company=extracted_row.get("company"),
                city=extracted_row.get("city"),
                state=extracted_row.get("state"),
                country=extracted_row.get("country"),
                role=extracted_row.get("role"),
                email=extracted_row.get("email"),
                website=extracted_row.get("website"),
                interest=extracted_row.get("interest"),
                notes=extracted_row.get("notes"),
                validation_status="invalid",
                validation_message="Pending validation.",
            )

            try:
                normalized_phone = self._normalize_campaign_phone(extracted_row.get("phone") or "")
                preview_row.normalized_phone = normalized_phone

                if normalized_phone in seen_phones:
                    preview_row.validation_status = "duplicate"
                    preview_row.validation_message = "Duplicate phone number found in this upload."
                elif not extracted_row.get("name"):
                    preview_row.validation_status = "invalid"
                    preview_row.validation_message = "The name field is required."
                else:
                    contact = CampaignContactInput(
                        name=extracted_row["name"],
                        phone=normalized_phone,
                        company=extracted_row.get("company"),
                        city=extracted_row.get("city"),
                        state=extracted_row.get("state"),
                        country=extracted_row.get("country"),
                        role=extracted_row.get("role"),
                        email=extracted_row.get("email"),
                        website=extracted_row.get("website"),
                        interest=extracted_row.get("interest"),
                        notes=extracted_row.get("notes"),
                        metadata=extracted_row.get("metadata") or {},
                    )
                    valid_contacts.append(contact)
                    seen_phones.add(normalized_phone)
                    preview_row.validation_status = "valid"
                    preview_row.validation_message = "Ready for campaign import."
            except ValueError as exc:
                preview_row.validation_status = "invalid"
                preview_row.validation_message = str(exc)

            preview_rows.append(preview_row)

        valid_count = len([row for row in preview_rows if row.validation_status == "valid"])
        invalid_count = len([row for row in preview_rows if row.validation_status == "invalid"])
        duplicate_count = len([row for row in preview_rows if row.validation_status == "duplicate"])

        return CampaignCsvPreviewResponse(
            filename=filename,
            file_type=file_type,
            total_rows=len(preview_rows),
            valid_contacts=valid_count,
            invalid_contacts=invalid_count,
            duplicate_contacts=duplicate_count,
            source_columns=source_columns,
            unmapped_columns=unmapped_columns,
            preview_rows=preview_rows,
            contacts=valid_contacts,
        )

    def _resolve_suffix(self, filename: str) -> str:
        lowered = filename.lower()
        for suffix in self.supported_file_types:
            if lowered.endswith(suffix):
                return suffix
        supported = ", ".join(sorted(self.supported_file_types))
        raise AppError(
            status_code=400,
            code="unsupported_lead_file_type",
            message=f"Supported renewal sheet file types are: {supported}.",
        )

    def _extract_rows(
        self,
        *,
        filename: str,
        suffix: str,
        raw_content: bytes,
    ) -> tuple[list[dict[str, object]], list[str], list[str]]:
        if suffix == ".csv":
            text = self._decode_text_bytes(raw_content)
            return self._extract_structured_rows_from_text(text)
        if suffix == ".xlsx":
            return self._extract_xlsx_rows(raw_content)
        if suffix == ".xls":
            return self._extract_xls_rows(raw_content)
        if suffix == ".docx":
            return self._extract_docx_rows(raw_content)
        if suffix == ".pdf":
            text = self._extract_pdf_text(raw_content)
            return self._extract_rows_from_text_document(text)
        if suffix == ".doc":
            text = self._extract_legacy_doc_text(raw_content)
            return self._extract_rows_from_text_document(text)
        if suffix == ".txt":
            text = self._decode_text_bytes(raw_content)
            return self._extract_rows_from_text_document(text)
        raise AppError(
            status_code=400,
            code="unsupported_lead_file_type",
            message=f"The lead file '{filename}' is not supported.",
        )

    def _extract_structured_rows_from_text(
        self,
        text: str,
    ) -> tuple[list[dict[str, object]], list[str], list[str]]:
        reader = csv.DictReader(io.StringIO(text))
        if reader.fieldnames is None:
            raise AppError(
                status_code=400,
                code="missing_lead_header",
                message="The uploaded spreadsheet must include a header row.",
            )

        headers = [self._clean_cell_value(header) for header in reader.fieldnames]
        rows = [list((row or {}).values()) for row in reader]
        return self._extract_structured_rows_from_matrix([headers, *rows])

    def _extract_xlsx_rows(self, raw_content: bytes) -> tuple[list[dict[str, object]], list[str], list[str]]:
        workbook = load_workbook(filename=io.BytesIO(raw_content), read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        matrix = [
            [self._clean_cell_value(cell) for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
        return self._extract_structured_rows_from_matrix(matrix)

    def _extract_xls_rows(self, raw_content: bytes) -> tuple[list[dict[str, object]], list[str], list[str]]:
        try:
            import xlrd
        except ModuleNotFoundError as exc:
            raise AppError(
                status_code=503,
                code="xls_support_missing",
                message="Legacy .xls support requires the 'xlrd' package to be installed.",
            ) from exc

        workbook = xlrd.open_workbook(file_contents=raw_content)
        sheet = workbook.sheet_by_index(0)
        matrix = [
            [self._clean_cell_value(sheet.cell_value(row_index, col_index)) for col_index in range(sheet.ncols)]
            for row_index in range(sheet.nrows)
        ]
        return self._extract_structured_rows_from_matrix(matrix)

    def _extract_docx_rows(self, raw_content: bytes) -> tuple[list[dict[str, object]], list[str], list[str]]:
        try:
            with zipfile.ZipFile(io.BytesIO(raw_content)) as archive:
                xml_content = archive.read("word/document.xml")
        except (KeyError, zipfile.BadZipFile) as exc:
            raise AppError(
                status_code=400,
                code="invalid_docx_file",
                message="The uploaded .docx file could not be read.",
            ) from exc

        root = ElementTree.fromstring(xml_content)
        namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}

        for table in root.findall(".//w:tbl", namespace):
            matrix: list[list[str | None]] = []
            for row in table.findall("./w:tr", namespace):
                values: list[str | None] = []
                for cell in row.findall("./w:tc", namespace):
                    texts = [node.text for node in cell.findall(".//w:t", namespace) if node.text]
                    values.append(" ".join(part.strip() for part in texts if part.strip()) or None)
                if any(values):
                    matrix.append(values)
            rows, source_columns, unmapped_columns = self._extract_structured_rows_from_matrix(matrix, raise_on_missing=False)
            if rows:
                return rows, source_columns, unmapped_columns

        paragraphs = []
        for paragraph in root.findall(".//w:p", namespace):
            texts = [node.text for node in paragraph.findall(".//w:t", namespace) if node.text]
            cleaned = " ".join(part.strip() for part in texts if part.strip()).strip()
            if cleaned:
                paragraphs.append(cleaned)
        return self._extract_rows_from_text_document("\n".join(paragraphs))

    def _extract_pdf_text(self, raw_content: bytes) -> str:
        try:
            reader = PdfReader(io.BytesIO(raw_content))
        except Exception as exc:  # pragma: no cover - parser exception details vary by input
            raise AppError(
                status_code=400,
                code="invalid_pdf_file",
                message="The uploaded PDF could not be read.",
            ) from exc

        pages = [page.extract_text() or "" for page in reader.pages]
        return "\n".join(page.strip() for page in pages if page and page.strip())

    def _extract_legacy_doc_text(self, raw_content: bytes) -> str:
        candidates: list[str] = []
        for encoding in ("utf-8", "utf-16", "latin-1"):
            try:
                candidates.append(raw_content.decode(encoding, errors="ignore"))
            except Exception:
                continue
        if not candidates:
            return ""

        best_candidate = max(candidates, key=lambda item: sum(char.isalpha() for char in item))
        best_candidate = best_candidate.replace("\x00", " ")
        best_candidate = re.sub(r"[\x01-\x08\x0B\x0C\x0E-\x1F]", " ", best_candidate)
        best_candidate = re.sub(r"[ \t]{2,}", " ", best_candidate)
        best_candidate = re.sub(r"\n{3,}", "\n\n", best_candidate)
        return best_candidate.strip()

    def _extract_rows_from_text_document(
        self,
        text: str,
    ) -> tuple[list[dict[str, object]], list[str], list[str]]:
        normalized = text.replace("\r", "\n").strip()
        if not normalized:
            return [], [], []

        lines = [line.strip() for line in normalized.splitlines() if line.strip()]
        for delimiter in ("\t", "|", ";", ","):
            matrix = [line.split(delimiter) for line in lines if delimiter in line]
            rows, source_columns, unmapped_columns = self._extract_structured_rows_from_matrix(matrix, raise_on_missing=False)
            if rows:
                return rows, source_columns, unmapped_columns

        blocks = [block.strip() for block in re.split(r"\n\s*\n", normalized) if block.strip()]
        extracted_rows: list[dict[str, object]] = []
        field_names: set[str] = set()
        for row_number, block in enumerate(blocks, start=1):
            record = self._extract_labeled_record(block)
            if not any(record.get(field) for field in self.supported_columns) and not record.get("metadata"):
                continue
            record["row_number"] = row_number
            extracted_rows.append(record)
            field_names.update(
                field_name
                for field_name in self.supported_columns
                if record.get(field_name)
            )
            field_names.update((record.get("metadata") or {}).keys())

        if extracted_rows:
            return extracted_rows, sorted(field_names), []
        return [], [], []

    def _extract_labeled_record(self, block: str) -> dict[str, object]:
        record: dict[str, object] = {column: None for column in self.supported_columns}
        metadata: dict[str, str] = {}
        for raw_line in block.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            matched_field = None
            for field_name, pattern in self.text_field_patterns.items():
                match = pattern.match(line)
                if not match:
                    continue
                record[field_name] = self._clean_cell_value(match.group(1))
                matched_field = field_name
                break
            if matched_field:
                continue

            generic_match = re.match(r"^([A-Za-z][A-Za-z0-9 /_&()]{1,40})\s*[:\-]\s*(.+)$", line)
            if generic_match:
                key = self._clean_cell_value(generic_match.group(1))
                value = self._clean_cell_value(generic_match.group(2))
                if key and value:
                    metadata[key] = value

        if metadata:
            record["metadata"] = metadata
        return record

    def _extract_structured_rows_from_matrix(
        self,
        matrix: list[list[object]],
        *,
        raise_on_missing: bool = True,
    ) -> tuple[list[dict[str, object]], list[str], list[str]]:
        trimmed_rows = [
            [self._clean_cell_value(cell) for cell in row]
            for row in matrix
            if any(self._clean_cell_value(cell) for cell in row)
        ]
        if not trimmed_rows:
            return [], [], []

        header_row = [cell or "" for cell in trimmed_rows[0]]
        mapped_headers = [self._match_canonical_field(header) for header in header_row]
        source_columns = [header for header in header_row if header]
        unmapped_columns = [
            header_row[index]
            for index, mapped_field in enumerate(mapped_headers)
            if header_row[index] and mapped_field is None
        ]
        if not self._has_required_contact_headers(mapped_headers):
            if raise_on_missing:
                raise AppError(
                    status_code=400,
                    code="missing_lead_columns",
                    message="The uploaded spreadsheet must contain phone details and either a contact name or company name.",
                )
            return [], source_columns, unmapped_columns

        structured_rows: list[dict[str, object]] = []
        for row_number, row in enumerate(trimmed_rows[1:], start=1):
            record: dict[str, object] = {column: None for column in self.supported_columns}
            metadata: dict[str, str] = {}
            for index, cell in enumerate(row):
                header = header_row[index] if index < len(header_row) else f"column_{index + 1}"
                if not header:
                    continue
                value = self._clean_cell_value(cell)
                if value is None:
                    continue
                mapped_field = mapped_headers[index] if index < len(mapped_headers) else None
                if mapped_field:
                    record[mapped_field] = value
                else:
                    metadata[header] = value
            if not any(record.get(column) for column in self.supported_columns) and not metadata:
                continue
            record = self._normalize_contact_identity(record)
            record["row_number"] = row_number
            record["metadata"] = metadata
            structured_rows.append(record)
        return structured_rows, source_columns, unmapped_columns

    @staticmethod
    def _has_required_contact_headers(mapped_headers: list[str | None]) -> bool:
        mapped_fields = {field for field in mapped_headers if field}
        return "phone" in mapped_fields and bool({"name", "company"} & mapped_fields)

    @staticmethod
    def _normalize_contact_identity(record: dict[str, object]) -> dict[str, object]:
        normalized_record = dict(record)
        if not normalized_record.get("name") and normalized_record.get("company"):
            normalized_record["name"] = normalized_record["company"]
        return normalized_record

    @staticmethod
    def _normalize_campaign_phone(phone_number: str) -> str:
        raw_value = phone_number.strip()
        digits_only = re.sub(r"\D", "", raw_value)
        if (
            raw_value
            and not raw_value.startswith("+")
            and not raw_value.startswith("00")
            and len(digits_only) == 10
            and digits_only[0] in {"6", "7", "8", "9"}
        ):
            return normalize_phone_number(f"+91{digits_only}")
        return normalize_phone_number(raw_value)

    def _decode_text_bytes(self, raw_content: bytes) -> str:
        for encoding in ("utf-8-sig", "utf-8", "utf-16", "latin-1"):
            try:
                return raw_content.decode(encoding)
            except UnicodeDecodeError:
                continue
        raise AppError(
            status_code=400,
            code="invalid_lead_encoding",
            message="The uploaded text file could not be decoded. Try UTF-8, UTF-16, or Latin-1 encoding.",
        )

    @staticmethod
    def _clean_cell_value(value: object) -> str | None:
        if value is None:
            return None
        if isinstance(value, float) and value.is_integer():
            text = str(int(value))
        else:
            text = str(value)
        cleaned = re.sub(r"\s+", " ", text).strip()
        return cleaned or None

    def _match_canonical_field(self, header: str | None) -> str | None:
        if not header:
            return None
        normalized = re.sub(r"[^a-z0-9]+", " ", header.lower()).strip()
        for canonical_field, aliases in self.header_aliases.items():
            if normalized in aliases:
                return canonical_field
        return None


@lru_cache
def get_campaign_csv_service() -> CampaignCsvService:
    return CampaignCsvService(get_settings())
